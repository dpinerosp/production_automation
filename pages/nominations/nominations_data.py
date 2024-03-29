import base64
from datetime import datetime
import io
import pandas as pd
from dash import html

from utils.functions import filter_data_by_date, load_companies, load_data, load_oil_types
from utils.constants import months, companies, nominations_data, balance_data

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def daily_transported_oil_type(data, start_date, end_date):
    """
    Return a DataFrame with the oils transported daily by oil type

    Parameters:
    -----------
    data: dataframe -> Dataframe with the production data

    Return:
    -------
    dataframe -> NSV transported daily by each oil type
    """

    filtered_data = filter_data_by_date(data, start_date, end_date)
    # Get just the transported oil of NSV
    transported = filtered_data[filtered_data['operacion'] == "DESPACHO POR REMITENTE"][['fecha', 'empresa', 'tipo crudo', 'NSV']]
    transported_oil_type = transported.pivot_table(values="NSV", 
                                                index=data["fecha"], 
                                                columns=["empresa", "tipo crudo"]
                                                ).reset_index()
    #transported_oil_type.reset_index(inplace=True)
    transported_oil_type["fecha"] = transported_oil_type['fecha'].dt.date
    transported_oil_type.set_index("fecha", inplace=True)
    transported_oil_type.fillna(0, inplace=True)
    return transported_oil_type

def get_date_nomination(filename):
    month_name = filename.split('_')[1].split('.')[0]
    month = months.index(month_name) + 1
    year = filename.split('_')[1].split('.')[1]
    start_date = datetime.strptime(f"01-{month}-{year}", "%d-%m-%Y")
    if month == 12:
        end_date = datetime.strptime(f"01-01-{year + 1}", "%d-%m-%Y")
    else:
        end_date = datetime.strptime(f"01-{month + 1}-{year}", "%d-%m-%Y")
    return (start_date, end_date)

def remove_entries_nominations(filepath, filename):
    (start_date, end_date) = get_date_nomination(filename)
    df = pd.read_csv(filepath)
    df['fecha'] = pd.to_datetime(df['fecha'], yearfirst=True)
    mask = (df['fecha'] < start_date) | (df['fecha'] >= end_date)
    filtered = df[mask]
    return filtered
    
def parse_contents(contents, filename, date, header):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    if 'xls' in filename:
        df = pd.read_excel(io.BytesIO(decoded), 
                            names=header, 
                            skiprows=4,
                            nrows=31).reset_index(drop=True)
        return df.dropna(how='all').fillna(0)
    return pd.DataFrame()

def filter_data_nominations(data, start_date, end_date, company):
    """Return a Dataframe filtered by period time and company"""
    filtered_by_date = filter_data_by_date(data, start_date, end_date)
    filter_columns = ['fecha'] + [column for column in filtered_by_date.columns if company.lower() in column.lower()]
    filtered_by_company = filtered_by_date[filter_columns]
    return filtered_by_company

def filter_data_transported(data, start_date, end_date, company):
    transported = daily_transported_oil_type(data, start_date, end_date)
    company_keys = {'geopark': 'geopark', 'verano': 'parex', 'parex': 'parex'}
    transported_by_company = transported[company_keys[company.lower()].upper()]
    transported_light_oil = calculate_transported_light_oil(transported_by_company)
    return transported_light_oil

def calculate_transported_light_oil(transported_data):
    oil_types = load_oil_types()
    data_light_oil = get_light_oil(transported_data)

    normal_oils = list(oil_types[oil_types['Livianos'] == 'NO']['Crudo'])
    normal_months = [column for column in transported_data.columns if column in normal_oils]
    data_normal_oils = transported_data[normal_months].reset_index()
    data_normal_oils['fecha'] = pd.to_datetime(data_normal_oils['fecha'], yearfirst=True)
    result = pd.concat([data_normal_oils, data_light_oil], axis = 1)
    return result.loc[:,~result.columns.duplicated()].copy()

def get_light_oil(transported_data):
    oil_types = load_oil_types()
    light_oils = list(oil_types[oil_types['Livianos'] == 'SI']['Crudo'])
    light_months = [column for column in transported_data.columns if column in light_oils]
    data_light_oil = transported_data[light_months].sum(axis=1)
    data_light_oil = data_light_oil.reset_index()
    data_light_oil['fecha'] = pd.to_datetime(data_light_oil['fecha'], yearfirst=True)
    data_light_oil.columns = ['fecha', 'Livianos']
    return data_light_oil

def get_data_percentage_nominations(start_date, end_date, type_graph = 'Tigana'):
    name_companies = load_companies()
    company_keys = {'geopark': 'geopark', 'parex': 'verano', 'verano': 'verano'}
    data = dict()
    for name_company in name_companies:
        aux = data_transported_nominated(start_date, end_date, company_keys[name_company.lower()])
        columns_nominations = ['fecha'] +  [column for column in aux[0].columns if type_graph.lower() in column.lower()]
        columns_transported = ['fecha'] + [column for column in aux[1].columns if type_graph.lower() in column.lower()]
        data[name_company] = (aux[0][columns_nominations], aux[1][columns_transported])
    return data
    
def data_transported_nominated(start_date, end_date, company):
    data_nominated = pd.read_csv(nominations_data)
    data_balance = load_data(balance_data)
    data_nominated['fecha'] = pd.to_datetime(data_nominated['fecha'], yearfirst=True)
    data_nominations = filter_data_nominations(data_nominated, start_date, end_date, company)
    data_transported = filter_data_transported(data_balance, start_date, end_date, company)
    return data_nominations, data_transported

def get_data_nominations_report(start_date, end_date):
    name_companies = load_companies()
    company_keys = {'geopark': 'geopark', 'parex': 'verano', 'verano': 'verano'}
    data = list()
    for name_company in name_companies:
        nominations, transported  = data_transported_nominated(start_date, end_date, company_keys[name_company.lower()])
        transported.columns = [column if company_keys[name_company.lower()] in column or column == 'fecha' else f"{column.lower()} {company_keys[name_company.lower()]}" for column in transported.columns]
        nominations.columns = [column if company_keys[name_company.lower()] in column or column == 'fecha' else f"{column.lower()} {company_keys[name_company.lower()]}" for column in nominations.columns]
        data.append(nominations.set_index('fecha'))
        data.append(transported.set_index('fecha'))
    df = pd.concat(data, axis=1)
    df = df.loc[:,~df.columns.duplicated()].copy().reset_index()
    df = df[['fecha', 'nominado jacana geopark', 'jacana estacion geopark',
            'nominado tigana geopark', 'tigana estacion geopark',
            'nominado livianos geopark','livianos geopark',
            'nominado jacana verano', 'jacana estacion verano',
           'nominado tigana verano',  'tigana estacion verano',
           'nominado cabrestero verano', 'cabrestero - bacano jacana estacion verano',
           'nominado livianos verano', 'livianos verano']]
    df['fecha'] = df['fecha'].dt.date
    return df

def styles_cell(cell, background_color, font_color):
    """
    Add style to indicated cell: background_color and font_color.
    """
    cell.fill = PatternFill('solid', fgColor=background_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="00000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell.font = Font(color=font_color, bold=True)

def add_styles_nominations(worksheet, background_color, font_color):
    for column in range(1, 16):
        for row in range(1, 35):
            cell = worksheet.cell(row=row, column=column)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == 1 or row > 32:
                thin = Side(border_style="thin", color="00000000")
                cell.font = Font(color=font_color, bold=True)
                worksheet.row_dimensions[row].height = 15
            else:
                cell.font = Font(color=font_color, bold=False)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        letter = get_column_letter(column)
        if column == 1:
            worksheet.column_dimensions[letter].width = 15
        else:
            worksheet.column_dimensions[letter].width = 13

def results_per_company(data_nominations_report, company_name, oils_company):
    columns_company = [column for column in data_nominations_report.columns if company_name.lower() in column]
    company_results = data_nominations_report[columns_company]
    columns_nominados_company = [column for column in company_results.columns if "nominado" in column]
    nominados_company_result = company_results[columns_nominados_company]
    columns_transported_company = [column for column in company_results.columns if "nominado" not in column]
    transported_company_result = company_results[columns_transported_company]

    result = pd.DataFrame()
    result["Remitente"] = [company_name]*len(oils_company)
    result["Crudos"] = oils_company
    result["Nominado Barriles NSV/Día promedio"] = [round(c, 2) for c in nominados_company_result.mean(axis=0, numeric_only=True).values]
    result["Transportado Barriles NSV/Día promedio"] = [round(c, 2) for c in transported_company_result.mean(axis=0, numeric_only=True).values]
    result["Notas"] = ""
    result["Factor de Servicio ODCA"] = ""
    return result